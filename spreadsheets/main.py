import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]



products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    product_number = product_list.cell(product_row, 1).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    inventory_price = product_list.cell(product_row, 5)

    # calculate the number of prodcust per Supplier
    if supplier_name in  products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculate total value of inventory per Supplier
    if supplier_name in  total_value_per_supplier:
        current_inv_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_inv_value + inventory * price
    else:
       total_value_per_supplier[supplier_name] = inventory * price

    # select products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_number)] = int(inventory)
    
    # add value for total inventory price

    inventory_price.value = '${:,.2f}'.format(inventory * price)

inventory_file.save("inventory_with_total_value.xlsx")

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
